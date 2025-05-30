from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from rest_framework import permissions, status, views
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import UserRegistrationSerializer, UserSerializer
from django.contrib.auth import authenticate, login
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.exceptions import AuthenticationFailed
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from .models import UploadedFile, ChatMessage, Course
from .forms import FileUploadForm, CourseForm
from django.contrib.auth import logout
from django.http import JsonResponse
import os
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
import json
from django.conf import settings
from storages.backends.gcloud import GoogleCloudStorage
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from google.cloud import storage
import datetime
from django.middleware.csrf import get_token
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.views.decorators.http import require_POST
import requests
import mimetypes
from io import BytesIO
import PyPDF2  
import docx    
from openpyxl import load_workbook  
import re
import io
import uuid
from django.core.files.base import ContentFile

import logging
logger = logging.getLogger(__name__)



# ============================================================
# INDEX VIEW
def index(request):
    """Landing page with Login and Register buttons."""
    return render(request, 'dashboard/index.html') 



# USER REGISTRATION VIEW
def register_view(request):
    """Registers a new user and redirects to login."""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password1')
        
        # Ensure username and email are unique
        if User.objects.filter(username=username).exists():
            return render(request, 'dashboard/register.html', {'error': 'Username already taken'})
        if User.objects.filter(email=email).exists():
            return render(request, 'dashboard/register.html', {'error': 'Email already in use'})

        # DEBUG: Check if User object is created
        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.save()
            print(f"User {username} registered successfully!")  # Debugging
            return redirect('login')  # Redirect to login page after registration
        except Exception as e:
            print(f"Registration failed: {e}")  # Print any exception that occurs
            return render(request, 'dashboard/register.html', {'error': 'Registration failed. Try again.'})

    return render(request, 'dashboard/register.html')




# JWT TOKEN GENERATION VIEW
class ObtainTokenPairView(views.APIView):
    """Authenticate user and generate JWT tokens."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        return Response(status=status.HTTP_401_UNAUTHORIZED)



# LOGIN VIEW
def login_view(request):
    """Logs in the user and redirects to their specific dashboard."""
    if request.user.is_authenticated:
        return redirect('user_dashboard', username=request.user.username)  # Redirect already logged-in users

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('user_dashboard', username=user.username)  # Redirects correctly to their own dashboard
        else:
            return render(request, 'dashboard/login.html', {'error': 'Invalid username or password'})

    return render(request, 'dashboard/login.html')


# LOGOUT VIEW
def logout_view(request):
    """Logs out the user and deletes all chatbot files."""
    if request.user.is_authenticated:
        try:
            # Delete all chatbot files first
            delete_chatbot_files(request)
        except Exception as e:
            logger.error(f"Error during logout cleanup: {str(e)}")
    
    # Then perform the logout
    logout(request)
    return redirect('login')



# USER-SPECIFIC DASHBOARD VIEW
@ensure_csrf_cookie
@login_required
def user_dashboard(request, username):
    """User-specific dashboard with courses."""
    if request.user.username != username:
        return redirect('login')  # Prevents unauthorized access

    # Get user's courses
    user_courses = Course.objects.filter(user=request.user)
    
    # Get selected course (from query param or first course or None)
    selected_course_id = request.GET.get('course')
    selected_course = None
    
    if selected_course_id:
        try:
            selected_course = Course.objects.get(id=selected_course_id, user=request.user)
        except Course.DoesNotExist:
            pass
    
    if not selected_course and user_courses.exists():
        selected_course = user_courses.first()
    
    # Get files filtered by course if selected_course exists
    if selected_course:
        user_files = UploadedFile.objects.filter(user=request.user, course=selected_course)
    else:
        # For backward compatibility, show files without course
        user_files = UploadedFile.objects.filter(user=request.user, course__isnull=True)
    
    # Forms
    file_form = FileUploadForm()
    course_form = CourseForm()

    # Handle file upload form submission
    if request.method == "POST":
        file_form = FileUploadForm(request.POST, request.FILES)
        if file_form.is_valid():
            try:
                uploaded_file = file_form.save(commit=False)
                uploaded_file.user = request.user
                
                # Associate with selected course if exists
                if selected_course:
                    uploaded_file.course = selected_course
                
                uploaded_file.save()
                print(f"File uploaded to: {uploaded_file.file.url}")
                
                # Redirect to the same page (as a GET request)
                return HttpResponseRedirect(request.path + (f'?course={selected_course.id}' if selected_course else ''))
            except Exception as e:
                print(f"Error uploading file: {str(e)}")
    
    # Render page
    context = {
        "files": user_files,
        "file_form": file_form,
        "course_form": course_form,
        "courses": user_courses,
        "selected_course": selected_course
    }
    
    response = render(request, 'dashboard/dashboard.html', context)
    response.set_cookie("csrftoken", get_token(request))
    return response


##### MULTIPLE COURSE FEATURE

# Add course management
@login_required
def add_course(request):
    """Add a new course for the user."""
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.user = request.user
            
            # Check for duplicate course name
            if Course.objects.filter(user=request.user, name=course.name).exists():
                return JsonResponse({"error": "You already have a course with this name"}, status=400)
            
            course.save()
            return JsonResponse({
                "id": course.id,
                "name": course.name,
                "created_at": course.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        else:
            return JsonResponse({"error": form.errors}, status=400)
    return JsonResponse({"error": "Only POST method is allowed"}, status=405)

@login_required
def delete_course(request, course_id):
    """Delete a course and all its files."""
    try:
        course = get_object_or_404(Course, id=course_id, user=request.user)
        
        # Get all files in this course
        files = UploadedFile.objects.filter(course=course)
        
        # Delete files from storage and database
        for file in files:
            file.file.delete()  # Delete from GCS
            file.delete()       # Delete from database
        
        # Delete the course
        course.delete()
        
        return JsonResponse({"message": "Course and all its files deleted successfully"}, status=200)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# SERVE FILE VIEW (Restricted to File Owner)
def serve_file(request, file_id):
    """Allow only the file owner to access their file."""
    try:
        # First check if user is authenticated via session
        if request.user.is_authenticated:
            user = request.user
        else:
            # Fall back to JWT if session auth is not available
            auth = JWTAuthentication()
            auth_user = auth.authenticate(request)
            if not auth_user:
                return redirect('login')
            user, _ = auth_user

        file = get_object_or_404(UploadedFile, id=file_id)
        
        # Check permissions
        if file.user != user:
            return HttpResponseForbidden("You do not have permission to access this file.")
        
        # Create a public URL that works without signing
        bucket_name = settings.GS_BUCKET_NAME
        file_path = file.file.name
        
        # Create the public URL
        public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
        
        # Redirect to the public URL
        return redirect(public_url)
        
    except AuthenticationFailed:
        return redirect('login')
    except UploadedFile.DoesNotExist:
        return HttpResponseForbidden("File not found.")




# DELETE FILE VIEW
@login_required
def delete_file_view(request, file_id):
    """Delete a file uploaded by the user."""
    try:
        # Authenticate user
        auth = JWTAuthentication()
        user_data = auth.authenticate(request)
        
        if user_data is None:
            # Fallback to session authentication
            user = request.user
        else:
            user, _ = user_data  
        
        # Retrieve the file or return 404 if it doesn't exist
        file = get_object_or_404(UploadedFile, id=file_id, user=user)

        # Delete file
        file.file.delete()  # from GCS
        file.delete()       # from database
        
        return JsonResponse({"message": "File deleted successfully"}, status=200)

    except AuthenticationFailed:
        return JsonResponse({"error": "Invalid credentials"}, status=403)
    except UploadedFile.DoesNotExist:
        return JsonResponse({"error": "File not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



#######################################################
## CHAT BOT FEATURE
# ChatBot
def get_groq_response(user_message):
    """Get a response from Groq API with enhanced system prompt."""
    try:
        headers = {
            "Authorization": f"Bearer {settings.GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        # Enhanced system prompt
        system_prompt = ("You are an educational assistant helping faculty with syllabus improvement, "
                        "question paper preparation, grading assessments, etc. "
                        "When asked to generate files, you MUST follow the file generation format instructions.")
        
        payload = {
            "model": "llama3-8b-8192",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_message}
            ],
            "temperature": 0.7,
            "max_tokens": 1024,
        }
        
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions", 
            headers=headers, 
            json=payload,
            timeout=30
        )
        
        if response.status_code != 200:
            return f"I'm sorry, there was an error with the AI service. Please try again later."
            
        return response.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return f"I'm sorry, I encountered an error: {str(e)}"


# Define allowed file types and size limits
ALLOWED_FILE_TYPES = {
    # Text files
    '.txt': 'text/plain',
    '.csv': 'text/csv',
    '.md': 'text/markdown',
    # Document files
    '.pdf': 'application/pdf',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    # Spreadsheets
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    # Presentations
    '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    # Add image files
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png'
}

# Maximum file size (5MB)
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB in bytes

# Maximum tokens for LLM
MAX_TOTAL_TOKENS = 7000
MAX_FILE_TOKENS = 5000

# Helper function for token estimation
def estimate_tokens(text):
    """Rough estimation of token count (4 chars ≈ 1 token)."""
    return len(text) // 4


# File content extraction
def extract_file_content(file_obj):
    """Extract text content from uploaded files."""
    try:
        # Get file extension
        file_name = file_obj.file.name
        file_ext = os.path.splitext(file_name)[1].lower()
        
        # Get file content
        file_content = file_obj.file.read()
        
        # Check file size again as a precaution
        if len(file_content) > MAX_FILE_SIZE:
            return "[File too large for processing]"

        # Image Files
        if file_ext in ['.jpg', '.jpeg', '.png']:
            return f"[This is an image file: {file_name}. The AI cannot directly see the image content, but can discuss it based on your description.]"
        
        # Text extraction based on file type
        if file_ext == '.txt' or file_ext == '.md' or file_ext == '.csv':
            # For text files, decode directly
            return file_content.decode('utf-8', errors='ignore')
            
        elif file_ext == '.pdf':
            try:
                # For PDFs, use PyPDF2
                pdf_reader = PyPDF2.PdfReader(BytesIO(file_content))
                text = ""
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
                return text
            except Exception as e:
                logger.error(f"Error extracting PDF content: {str(e)}")
                return "[Error extracting PDF content]"
                
        elif file_ext == '.docx':
            try:
                # For DOCX, use python-docx
                doc = docx.Document(BytesIO(file_content))
                text = "\n".join([para.text for para in doc.paragraphs if para.text])
                return text
            except Exception as e:
                logger.error(f"Error extracting DOCX content: {str(e)}")
                return "[Error extracting DOCX content]"
                
        elif file_ext == '.xlsx':
            try:
                # For Excel files, use openpyxl
                workbook = load_workbook(BytesIO(file_content), read_only=True)
                text = ""
                # Get the first few sheets
                for sheet_name in workbook.sheetnames[:3]:  # Limit to first 3 sheets
                    sheet = workbook[sheet_name]
                    text += f"Sheet: {sheet_name}\n"
                    # Get first 100 rows and 20 columns max
                    for row in list(sheet.rows)[:100]:
                        row_values = [str(cell.value) if cell.value is not None else "" for cell in row[:20]]
                        text += ", ".join(row_values) + "\n"
                    text += "\n"
                return text
            except Exception as e:
                logger.error(f"Error extracting Excel content: {str(e)}")
                return "[Error extracting Excel content]"
                
        elif file_ext == '.pptx':
            # For PPTX, more complex extraction would be needed
            # Using a placeholder for now
            return "[This is a PowerPoint file. Text extraction requires additional processing.]"
            
        else:
            return f"[File type {file_ext} not supported for content extraction]"
            
    except Exception as e:
        logger.error(f"Error extracting content from file: {str(e)}")
        return "[Error processing file]"
    finally:
        # Reset file pointer position
        if hasattr(file_obj.file, 'seek'):
            file_obj.file.seek(0)

# Handle file uploads for chat
@csrf_exempt
@login_required
def upload_chat_file(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Log detailed file info
        file_name = uploaded_file.name
        file_ext = os.path.splitext(file_name)[1].lower()
        file_content_type = uploaded_file.content_type
        file_size = uploaded_file.size
        
        logger.info(f"File upload attempt: name={file_name}, ext={file_ext}, type={file_content_type}, size={file_size}")
        
        # Check file size
        if file_size > MAX_FILE_SIZE:
            logger.warning(f"File size too large: {file_size} bytes > {MAX_FILE_SIZE} bytes")
            return JsonResponse({
                'error': f'File size exceeds the maximum limit of 5MB'
            }, status=400)
        
        # Check file type
        if file_ext not in ALLOWED_FILE_TYPES:
            logger.warning(f"File type not allowed: {file_ext}")
            return JsonResponse({
                'error': f'File type {file_ext} is not supported. Supported types: {", ".join(ALLOWED_FILE_TYPES.keys())}'
            }, status=400)
        
        # Specifically handle PNG files
        if file_ext == '.png':
            logger.info(f"PNG file detected: {file_name}")
            # Check if content type is incorrect
            if not file_content_type.startswith('image/'):
                logger.info(f"Fixing content type for PNG: {file_content_type} -> image/png")
                # Force the correct content type for PNG files
                uploaded_file.content_type = 'image/png'
        
        # Use a custom form for better validation
        form = FileUploadForm({'category': 'chatbot_files'}, {'file': uploaded_file})
        
        # Check if form is valid
        if form.is_valid():
            logger.info(f"Form is valid for file: {file_name}")
            file_obj = form.save(commit=False)
            file_obj.user = request.user
            file_obj.save()
            
            # Create a public URL manually
            bucket_name = settings.GS_BUCKET_NAME
            file_path = file_obj.file.name
            public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
            
            response_data = {
                'file_id': file_obj.id,
                'file_name': file_obj.file.name,
                'file_url': public_url
            }
            logger.info(f"File uploaded successfully: {response_data}")
            return JsonResponse(response_data)
        else:
            logger.error(f"Form validation failed: {form.errors}")
            return JsonResponse({'error': f'Form validation error: {form.errors}'}, status=400)
    
    except Exception as e:
        logger.error(f"Exception in upload_chat_file: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=500)


# Chatbot File Generation feature
@login_required
def generate_file(request):
    """Generate a file from chatbot content and save it to storage."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        content = data.get("content", "")
        file_name = data.get("file_name", "")
        file_type = data.get("file_type", "txt")
        
        if not content or not file_name:
            return JsonResponse({'error': 'Content and filename are required'}, status=400)
        
        # Validate file type
        valid_types = ['txt', 'md', 'csv']
        if file_type not in valid_types:
            file_type = 'txt'  # Default to txt if invalid
            
        # Ensure file has proper extension
        if not file_name.endswith(f'.{file_type}'):
            file_name = f"{file_name}.{file_type}"
        
        # Create file content
        file_content = ContentFile(content.encode('utf-8'))
        
        # Save to UploadedFile model with chatbot_files category
        file_obj = UploadedFile(
            user=request.user,
            category="chatbot_files"
        )
        
        # Associate with course if selected
        selected_course_id = data.get('course_id')
        if selected_course_id:
            try:
                course = Course.objects.get(id=selected_course_id, user=request.user)
                file_obj.course = course
            except Course.DoesNotExist:
                pass
        
        # Save the file to storage
        file_obj.file.save(file_name, file_content)
        file_obj.save()
        
        # Create a public URL manually
        bucket_name = settings.GS_BUCKET_NAME
        file_path = file_obj.file.name
        public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
        
        return JsonResponse({
            'success': True,
            'file_id': file_obj.id,
            'file_name': file_obj.file.name,
            'file_url': public_url
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error generating file: {str(e)}")
        return JsonResponse({'error': f'Error generating file: {str(e)}'}, status=500)


# ChatBot API
@csrf_exempt
@require_POST
def chatbot_api(request):
    """API endpoint for chatbot interactions with file support and generation."""
    try:
        data = json.loads(request.body)
        user_message = data.get("message", "")
        file_ids = data.get("file_ids", [])
        
        if not user_message:
            return JsonResponse({"error": "Empty message"}, status=400)
        
        # Get files content for context if file_ids are provided
        file_context = ""
        associated_files = []
        total_tokens = estimate_tokens(user_message)
        
        if file_ids and request.user.is_authenticated:
            for file_id in file_ids:
                try:
                    file_obj = UploadedFile.objects.get(id=file_id, user=request.user)
                    associated_files.append(file_obj)
                    
                    # Extract text content from the file
                    file_content = extract_file_content(file_obj)
                    
                    if file_content:
                        # Estimate tokens for this file content
                        file_tokens = estimate_tokens(file_content)
                        
                        # Check if adding this would exceed our limit
                        if total_tokens + file_tokens > MAX_TOTAL_TOKENS:
                            # Truncate content to fit within limits if possible
                            if file_tokens > 100:  # Only truncate if it's worth it
                                max_chars = (MAX_TOTAL_TOKENS - total_tokens) * 4
                                file_content = file_content[:max_chars] + "...[content truncated due to length]"
                                file_tokens = estimate_tokens(file_content)
                            
                        # Update total tokens and add to context
                        total_tokens += file_tokens
                        file_context += f"\nContent from file '{file_obj.file.name}':\n{file_content}\n"
                        
                        # If we've used up our token budget, stop processing files
                        if total_tokens >= MAX_FILE_TOKENS:
                            file_context += "\n[Additional files were not processed due to length constraints]"
                            break
                            
                except UploadedFile.DoesNotExist:
                    continue
        
        # More explicit file generation detection
        should_generate_file = any(keyword in user_message.lower() for keyword in 
            ['generate file', 'create file', 'make a file', 'save as file', 'generate a file', 
             'create a file', 'make file', 'save file', 'download file', 'file with'])
        
        # Enhanced file generation prompt
        generate_file_prompt = ""
        if should_generate_file:
            generate_file_prompt = (
                "\nIMPORTANT: The user is asking you to generate a file. "
                "You MUST generate this file by following these instructions exactly:\n"
                "1. In your response, first acknowledge that you'll create the file\n"
                "2. Then add the file content between these exact markers: [FILE_CONTENT_START] and [FILE_CONTENT_END]\n"
                "3. Also include 'Filename: your-suggested-filename.txt' on its own line\n"
                "4. Only use .txt, .md, or .csv file extensions as these are the only supported formats\n"
                "EXAMPLE RESPONSE FORMAT:\n"
                "I'll generate that file for you!\n\n"
                "Filename: grading_rubric.txt\n\n"
                "[FILE_CONTENT_START]\n"
                "Your file content here...\n"
                "[FILE_CONTENT_END]\n\n"
                "Here's your file! You can download it using the link below."
            )
        
        # Enhanced system prompt for the AI
        system_prompt = ("You are an educational assistant helping faculty with syllabus improvement, "
                         "question paper preparation, grading assessments, etc. "
                         "When asked to generate files, you MUST follow the file generation format instructions.")
        
        # Prepare enhanced prompt with file context
        enhanced_message = user_message
        if file_context:
            enhanced_message = f"The user asked: {user_message}\n\nI'm also providing content from files they uploaded to help answer their question:{file_context}\n\nBased on this information, please respond to their question.{generate_file_prompt}"
        else:
            enhanced_message = f"The user asked: {user_message}\n\nPlease respond to their question.{generate_file_prompt}"
        
        # Update the request payload to include our enhanced system prompt
        headers = {
            "Authorization": f"Bearer {settings.GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "llama3-8b-8192",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": enhanced_message}
            ],
            "temperature": 0.7,
            "max_tokens": 1024,
        }
        
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions", 
            headers=headers, 
            json=payload,
            timeout=30
        )
        
        if response.status_code != 200:
            return JsonResponse({"response": "I'm sorry, there was an error with the AI service. Please try again later."})
            
        response_text = response.json()["choices"][0]["message"]["content"]
        
        # Process file generation markers
        file_generation_data = None
        if "[FILE_CONTENT_START]" in response_text and "[FILE_CONTENT_END]" in response_text:
            # File content extraction logic
            file_start = response_text.find("[FILE_CONTENT_START]") + len("[FILE_CONTENT_START]")
            file_end = response_text.find("[FILE_CONTENT_END]")
            file_content = response_text[file_start:file_end].strip()
            
            # Remove markers from response
            original_response = response_text
            response_text = response_text.replace(
                response_text[response_text.find("[FILE_CONTENT_START]"):file_end + len("[FILE_CONTENT_END]")], 
                "(File content below)"
            )
            
            # Extract filename with proper validation
            filename = "generated_file.txt"  # Default
            if "filename:" in original_response.lower():
                filename_pattern = re.compile(r'filename:[\s]*([^\n]+)', re.IGNORECASE)
                filename_match = filename_pattern.search(original_response)
                if filename_match:
                    suggested_filename = filename_match.group(1).strip()
                    # Ensure the file has a text-based extension
                    valid_extensions = ['.txt', '.md', '.csv']
                    has_valid_extension = False
                    for ext in valid_extensions:
                        if suggested_filename.lower().endswith(ext):
                            has_valid_extension = True
                            filename = suggested_filename
                            break
                    
                    if not has_valid_extension:
                        # If no valid extension, add .txt
                        filename = suggested_filename + '.txt'
            
            # Create file generation data
            file_generation_data = {
                "content": file_content,
                "filename": filename
            }
            
            # Also create the file immediately for the user
            try:
                file_content_bytes = file_content.encode('utf-8')
                file_obj = ContentFile(file_content_bytes)
                
                # Create UploadedFile entry
                uploaded_file = UploadedFile(
                    user=request.user,
                    category="chatbot_files"
                )
                
                # Associate with course if selected
                if request.user.is_authenticated and 'course_id' in data:
                    try:
                        course = Course.objects.get(id=data['course_id'], user=request.user)
                        uploaded_file.course = course
                    except Course.DoesNotExist:
                        pass
                
                # Save file to storage
                uploaded_file.file.save(filename, file_obj)
                uploaded_file.save()
                
                # Create a public URL
                bucket_name = settings.GS_BUCKET_NAME
                file_path = uploaded_file.file.name
                public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
                
                # Update file generation data with file info
                file_generation_data.update({
                    "file_id": uploaded_file.id,
                    "file_url": public_url
                })
                
            except Exception as e:
                logger.error(f"Error saving generated file: {str(e)}")
                # Continue even if file saving fails
        
        # Save to database
        if request.user.is_authenticated:
            try:
                chat_msg = ChatMessage.objects.create(
                    user=request.user,
                    message=user_message,
                    response=response_text
                )
                # Associate files with this chat message
                if associated_files:
                    chat_msg.files.add(*associated_files)
            except Exception as e:
                logger.error(f"Error saving chat message: {str(e)}")
                # Continue even if saving fails
                pass
        
        # Return response
        response_data = {"response": response_text}
        if file_generation_data:
            response_data["file_generation"] = file_generation_data
            
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Error in chatbot API: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


# Retrive chat history
@login_required
def chat_history(request):
    """Retrieve chat history for the current user."""
    chats = ChatMessage.objects.filter(user=request.user)
    
    # Convert to list of dicts for JSON response
    chat_list = [
        {
            'id': chat.id,
            'message': chat.message,
            'response': chat.response,
            'timestamp': chat.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }
        for chat in chats
    ]
    
    return JsonResponse({'chats': chat_list})

# Delete Chatbot files
@login_required
def delete_chatbot_files(request):
    """Delete all chatbot files (uploaded and generated) for the current user."""
    try:
        # Get all chat messages for this user
        chat_messages = ChatMessage.objects.filter(user=request.user)
        
        # Get all files associated with these chat messages
        chatbot_uploaded_files = set()
        for chat in chat_messages:
            # Get files associated with this chat message
            for file in chat.files.all():
                chatbot_uploaded_files.add(file.id)
        
        # Also include any files specifically categorized as chatbot_files
        chatbot_generated_files = UploadedFile.objects.filter(
            user=request.user,
            category="chatbot_files"
        ).values_list('id', flat=True)
        
        # Combine both sets of files
        all_chatbot_files_ids = list(chatbot_uploaded_files) + list(chatbot_generated_files)
        all_chatbot_files = UploadedFile.objects.filter(id__in=all_chatbot_files_ids)
        
        # Log how many files will be deleted
        file_count = all_chatbot_files.count()
        logger.info(f"Deleting {file_count} chatbot files for user {request.user.username}")
        
        # Delete each file from storage and database
        for file in all_chatbot_files:
            try:
                # Delete from GCS
                file.file.delete()
                # Delete from database
                file.delete()
            except Exception as e:
                logger.error(f"Error deleting file {file.id}: {str(e)}")
        
        # Clear file associations from chat messages
        for chat in chat_messages:
            chat.files.clear()
        
        return JsonResponse({"success": True, "count": file_count})
    
    except Exception as e:
        logger.error(f"Error deleting chatbot files: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)